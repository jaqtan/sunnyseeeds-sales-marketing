import requests
import openpyxl
from datetime import datetime

# Get API key from environment
import os
api_key = os.getenv('TYPEFORM_API_KEY')
form_id = os.getenv('TYPEFORM_FORM_ID')

if not api_key or not form_id:
    print("ERROR: TYPEFORM_API_KEY or TYPEFORM_FORM_ID not set!")
    exit(2)

# Fetch from Typeform
url = f"https://api.typeform.com/forms/{form_id}/responses"
headers = {"Authorization": f"Bearer {api_key}"}

try:
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    data = response.json()
    
    # Count responses by program
    responses = data.get('items', [])
    print(f"Total Typeform responses: {len(responses)}")
    
    # Load Excel
    excel_path = "dashboards/SunnySeeds_Sales_Funnel_Dashboard.xlsx"
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Update timestamp
    ws['A2'] = f"Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Update Row 8 with total lead count
    # For now, just put total responses in column B (Preschool)
    ws['B8'] = len(responses)
    
    # Save
    wb.save(excel_path)
    print(f"✅ Updated Excel with {len(responses)} leads")
    
except Exception as e:
    print(f"ERROR: {str(e)}")
    exit(2)
